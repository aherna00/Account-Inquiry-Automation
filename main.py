from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time

fidelity_accounts = []

read_acc = load_workbook(filename="yoursheethere.xlsx")

sheet = read_acc.active
m_row = sheet.max_row

# pull all account numbers in excel into a list

for i in range(1, m_row + 1):
    cell_acc = sheet.cell(row=i + 1, column=1)
    fidelity_accounts.append(cell_acc.value)
print(fidelity_accounts)


# check all accounts for funded balance

def search(pending):
    for i in pending:
        driver.find_element_by_xpath \
                (
                "//ui-view/fip-basis-layout-portal-container/div/div/div/fip-side-bar-menu/bss-side-menu/div[1]/button").click()
        look_up = driver.find_element_by_id("ual-search-primary-input-unmasked")
        look_up.clear()
        look_up.send_keys(i)
        time.sleep(1)
        search_click = driver.find_element_by_xpath \
                (
                "/html/body/div/ui-view/fip-basis-layout-portal-container/div/div/div/fip-side-bar-item-panel/div[1]/div[2]/fip-unified-access-layer/unified-access-layer/div/unified-access-layer/div/div[1]/unified-access-layer-search-bar-sidebar/div/div/div/div/div/div[1]/div/span[5]/button")
        search_click.click()
        time.sleep(3)
        balance_button = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.ID, "account-cartridge__block_balanceinfo")))
        balance_button.click()
        balance = driver.find_element_by_xpath(
            "/html/body/div/ui-view/fip-basis-layout-portal-container/div/div/div/main/ui-view[2]/ui-view/fip-basis-layout-account-page/div/div[1]/account-cartridge/div/div[2]/div/extend-cartridge-details/section[2]/div[1]/div[1]/ul/li[1]/div/span").get_attribute(
            "innerHTML")
        if balance != "$0.00":
            print(i + " is funded " + balance)
        else:
            print("")


driver = webdriver.Chrome()
driver.implicitly_wait(60)
driver.get("https://www.wealthscape.com/")
driver.maximize_window()
main_window = driver.window_handles[0]

user_name = driver.find_element_by_id("userInput")
user_name.send_keys("user")

driver.find_element_by_class_name("group-h__item-spaced").click()

password = driver.find_element_by_id("password")

password.send_keys("pass")
driver.find_element_by_id("fs-login-button").click()

search(fidelity_accounts)

driver.close()
